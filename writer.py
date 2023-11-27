import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.colors import Color
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


def save_schedule(individual, path, kwargs):
    wb = openpyxl.Workbook() 
    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    columns=['A', 'B', 'C', 'D', 'E', 'F']
    hour_size = kwargs['hour_size']
    course_of_studying = individual.candidate['COURSE OF STUDYING'].unique() 

    # create a sheet for each course of studying
    for page, indirizzo in enumerate(course_of_studying):
        wb.create_sheet(index= page, title=indirizzo) 
        sheet = wb[indirizzo]
        df = individual.candidate[individual.candidate['COURSE OF STUDYING'] == indirizzo] 
        df = df.reset_index(drop=True)
        corsi = df['TEACHING'].unique().tolist() 
        colors = []

        # define a color for each teaching
        seed = np.random.randint(44, 56)
        for i in range(len(corsi)): 
            color = Color(indexed = seed + i)
            colors.append(color)

        # define schedule columns (days)
        for i in range(len(days)): 
            c = sheet.cell(row = 1, column = i+2) 
            c.value =  days[i]
            c.font = Font(size=11, bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')


        # define schedule rows (hours)
        for i in range(hour_size): 
            c = sheet.cell(row = i+2, column = 1) 
            c.value =  str(8+i)+":30 - "+str(9+i)+":30"
            c.font = Font(size=11, bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')

        # insert cells' values (teaching and room)
        for i in range(len(df)): 
            for hour in df.at[i, 'HOUR']:
                c = sheet.cell(row = hour+1, column = df.at[i, 'DAY']+1)
                color_index = corsi.index(df.at[i, 'TEACHING'])
                cell_color = colors[color_index]
                value_tmp = df.at[i, 'TEACHING']+'(ROOM '+str(df.at[i, 'ROOM'])+')'
                if c.value is not None: 
                    old_value = c.value 
                    if ':' in old_value: c.value =  str(old_value)+',\n'+value_tmp
                    else: c.value =  'OVERLAP:\n'+str(old_value)+',\n'+value_tmp
                    cell_color = '00FF0000' # if there is an overlap, cell_color = red
                else: c.value =  value_tmp
                c.fill = PatternFill(start_color=cell_color, fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center')

        # set columns width
        for i in range(len(columns)): sheet.column_dimensions[columns[i]].width = 30 

        # # set rows height
        # for i in range(hour_size): sheet.row_dimensions[i+2].height = 70 

        if individual.room_overlap is not None:
            for overlap in individual.room_overlap:
                day = overlap['DAY']
                hour = overlap['HOUR']
                room = overlap['ROOM']
                c = sheet.cell(row = hour+1, column = day+1) 
                if c.value is not None:
                    if str(room) in c.value: 
                        c.fill = PatternFill(start_color='00FF0000', fill_type='solid')


    del wb['Sheet'] # remove the defaul sheet
    wb.save(path)
    print('Schedule saved in:\t', path)



def save_population(population, path):
    with pd.ExcelWriter(path) as writer:
        for i, individual in enumerate(population): 
            individual.candidate.to_excel(writer, sheet_name="Schedule_"+str(i))
    print('Population saved in:\t', path)



def save_individual(individual, path):
    candidate = individual.candidate
    with pd.ExcelWriter(path) as writer:
        candidate.to_excel(writer)
    print('Individual saved in:\t', path)


def save_fitness(fitness_list, path):
    max_fitness = []
    min_fitness = []
    mean_fitness = []
    for generattion_fitness in fitness_list:
        generattion_fitness.sort()
        max_fitness.append(generattion_fitness[-1])
        min_fitness.append(generattion_fitness[0])
        mean_fitness.append(sum(generattion_fitness)/len(generattion_fitness))
    
    plt.figure()
    plt.title("FITNESS")
    plt.plot(max_fitness, color="red", linewidth=3, label="Worst")
    plt.plot(min_fitness, color="lime", linewidth=3, label="Best")
    plt.plot(mean_fitness, color="orange", linewidth=3, label="Mean")
    plt.legend()
    plt.savefig(path)
    plt.close()
    print('Fitness saved in:\t', path)


def save_diversity(diversity_list, path):
    plt.figure()
    plt.title("DIVERSITY")
    plt.plot(diversity_list, color="orange", linewidth=3)
    plt.savefig(path)
    plt.close()
    print('Diversity saved in:\t', path)


def save_statistics(best_individual, kwargs, path):
    with open(path, 'w') as f:
        f.write('OPTIMIZATION PARAMETERS:\n')
        for k, v in kwargs.items():
            f.write(k+' = '+str(v)+'\n')
        f.write('\n\nOPTIMIZATION STATISTICS:\n')
        f.write('Best Fitness: '+str(best_individual.fitness))
        f.write('\nNumber of rooms overlapped: '+str(best_individual.room_overlap))
        f.write('\nNumber of lessons overlapped: '+str(best_individual.hour_overlap))
        f.write('\nNumber of mandatory teachings overlapped: '+str(best_individual.type_overlap))
        
    print('Statistics saved in:\t', path)
